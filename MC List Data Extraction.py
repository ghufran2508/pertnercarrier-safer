import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import re
import json
import os
import multiprocessing
from playsound import playsound

def play_sound_loop(sound_file):
    while True:
        playsound(sound_file)

class WebScrapper:
    def __init__(self):
        # CONSTANTS
        # headers required for bs4
        self.__HEADERS__ = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
        self.__PARTNERCARRIES_URL__ = 'https://partnercarrier.com/'
        self.__SAFERSCRAPE_URL__ = 'https://safer.fmcsa.dot.gov/query.asp'
        self.__SHEET_HEADERS__ = ["MC Number", "US Dot", "Company Name", "Phone Number", "Email", "Physical Address", "Number of Trucks", "Authority Status"]

        # creating session
        # -> required for saferscraping
        self.session = requests.Session()

        # workbook
        self.__file_name__ = ""
        self.__workbook__ = None
        self.__worksheet__ = None
        self.__lastSheet__ = None

    def __del__(self):
        if self.__workbook__ is not None:
            self.__workbook__.close()
    
    # check if a partnercarries contains company input by user. 
    def __check_if_company_exist__(self, state_name):
        response = self.session.get(self.__PARTNERCARRIES_URL__, headers=self.__HEADERS__)

        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')

            links = soup.find_all("a", class_="link-font-size", href=True)

            for companies in links:
                # links contain data in format "Georgia Trucking Companies (1234)"
                # this line will only check company name in "georgia trucking companies"
                if state_name in re.sub('([0-9])', '', companies.text).lower():
                    print(f"Company {state_name} found in partnercarries. URL {companies.get('href')}")
                    return companies.get('href')

        else:
            print("Error: Failed to retrieve data")
            print(response.status_code)

        return None
        
    def __get_all_cities__(self, city):
        city_url = self.__PARTNERCARRIES_URL__+city[1::]
        list_of_all_cities_url = []
        print(city_url)
        response = self.session.get(city_url, headers=self.__HEADERS__)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')

            ALPHABETS = [chr(x+65) for x in range(0, 26)]

            for alpha in ALPHABETS:
                cities_container = soup.find('h4', class_="text-center h2backcolor well", id=alpha)
                if cities_container is None:
                    continue
                cities_container = cities_container.parent
                cities_name = cities_container.find_all('a', class_="city-link-font-size", href=True)
                print(f"{len(cities_name)} cities found in {alpha}...")
                
                list_of_cities_url = [city.get('href') for city in cities_name]
                list_of_all_cities_url.append(list_of_cities_url)
        else:
            print("Error: Failed to retrieve data")
            print(response.status_code)

        return list_of_all_cities_url

    def __get_list_of_companiesVM__(self, page_url):
        custom_header = self.__HEADERS__
        custom_header['Content-type'] = 'application/x-www-form-urlencoded'
        custom_header['X-Requested-With'] = 'XMLHttpRequest'

        has_next_page = True
        page_index = 1
        mc_list = []

        while has_next_page:
            ENTITY = '/entity-C'; query = '/truckortractor-1-5,6-10,11-20/'; queryPage = f"?p={page_index}&se=LegalName&sd=Asc" if page_index > 1 else ""
            custom_query = ENTITY+query+queryPage
            new_page_url = self.__PARTNERCARRIES_URL__+page_url[1::]+custom_query
            page_index += 1

            response = self.session.post(new_page_url, headers=custom_header)
            if response.status_code == 200:
                try:
                    dict = response.json()
                    has_next_page = dict["companyVM"]["Companies"]["Pagination"]["HasNextPage"]
                    all_companies = dict["companyVM"]["Companies"]["Items"]

                    for dataDict in all_companies:
                        mc_number = dataDict["IccDocketNumberFirst"]
                        if mc_number:
                            mc_list.append(mc_number)

                except Exception:
                    print("data loading failed")
                    has_next_page = False

            else:
                print("Error: Failed to retrieve data")
                print(response.status_code)
        
        return mc_list
    
    def __get_safer_data__(self, mcn):
        form_data = {'searchtype': 'ANY', 'query_type': 'queryCarrierSnapshot', 'query_param': 'MC_MX', 'query_string': mcn}

        session = requests.Session()
        response = session.post(self.__SAFERSCRAPE_URL__, headers=self.__HEADERS__, data=form_data)

        print(response)
        if response.status_code == 200:
            soup = BeautifulSoup(response.content, 'html.parser')

            trs = soup.find_all('tr')

            target_trs = [tr for tr in trs if tr.find('th', class_='querylabelbkg')]

            if len(target_trs) == 0:
                print("No Data found")
                return None
            else:
                if target_trs[0].find('td', class_='queryfield').text.strip() == "CARRIER":
                    if target_trs[1].find('td', class_='queryfield').text.strip() == "ACTIVE":
                        usdot = target_trs[2].find('td', class_='queryfield').text.strip()
                        authority_status = target_trs[4].find('td', class_='queryfield').text.split('  ')[0]

                        if authority_status in ["NOT FOUND", "NOT AUTHORIZED", "OUT OF SERVICE", "AUTHORIZED FOR BROKER Property", "AUTHORIZED FOR Passenger", "AUTHORIZED FOR HHG", "\nNOT AUTHORIZED \n\n\n*Please Note:\nNOT AUTHORIZED does not apply to Private or Intrastate operations.\n\nFor Licensing and Insurance details"]:
                            print("No Data Found")
                            return None

                        state_name = ' '.join([word.capitalize() for word in target_trs[6].find('td', class_='queryfield').text.strip().split()])
                        address = target_trs[8].find('td', class_='queryfield').text.strip()
                        address_parts = address.split('\n')
                        if len(address_parts) >= 2:
                            street_address = ' '.join([part.capitalize() for part in address_parts[0].split(' ')])
                            city_state_zip_parts = [part.strip().capitalize() for part in address_parts[1].split(',')]
                            city_state_zip_parts[0] = ' '.join(
                                [word.capitalize() for word in city_state_zip_parts[0].split(' ')])
                            city_state_zip_parts[-1] = city_state_zip_parts[-1].upper()
                            city_state_zip = ', '.join(city_state_zip_parts)

                            # Extract state abbreviation
                            state_abbreviation = city_state_zip_parts[1].split(' ')[0].strip()

                            # Check if state is in contiguous US
                            if state_abbreviation in [
                                'AL', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA', 'ID', 'IL', 'IN', 'IA',
                                'KS', 'KY', 'LA', 'ME', 'MD', 'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV',
                                'NH', 'NJ', 'NM', 'NY', 'NC', 'ND', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD',
                                'TN', 'TX', 'UT', 'VT', 'VA', 'WA', 'WV', 'WI', 'WY'
                            ]:
                                address = f"{street_address}\n{city_state_zip}"
                            else:
                                print(f"State {state_abbreviation} is not in the contiguous US. Skipping.")
                                return None

                        else:
                            print(f"Error: Address format is unexpected for MC number {mcn}. Address: {address}")

                        phone = target_trs[9].find('td', class_='queryfield').text.strip()
                        number_of_trucks = 'NOT FOUND'
                        for tr in target_trs:
                            if 'Power Units' in tr.text:
                                number_of_trucks = tr.find('td', class_='queryfield').text.strip()
                                break
                        
                        try:
                            number_of_trucks_int = int(number_of_trucks)
                            if number_of_trucks_int < 1 or number_of_trucks_int > 20:
                                print("Carrier does not meet truck count criteria")
                                return None
                        except ValueError:
                            print("Invalid number of trucks")
                            return None
                        
                        # Check Cargo Carried section
                        tables = soup.find('table', summary="Cargo Carried")
                        flags_index = [7, 15, 18, 19, 22, 31]
                        trs_cargo_carried = tables.find_all('tr')

                        for flags in flags_index:
                            row_cargo_carried = trs_cargo_carried[flags]
                            if row_cargo_carried.find('td', class_='queryfield').text.strip() == "X":
                                print(f"Carrier is marked with {row_cargo_carried} in Cargo Carried section with an X for MC number {mcn}")
                                return None

                        email = 'NOT FOUND'
                        url2 = f'https://ai.fmcsa.dot.gov/SMS/Carrier/{usdot}/CarrierRegistration.aspx'
                        response2 = session.get(url2, headers=self.__HEADERS__)

                        if response2.status_code == 200:
                            soup2 = BeautifulSoup(response2.content, 'html.parser')
                            email_div = soup2.find('div', class_='modalBody')

                            if email_div:
                                spans = email_div.findAll('span', class_='dat')
                                if len(spans) > 7:
                                    email = spans[6].text.strip().lower()

                        return usdot, state_name, address, phone, email, number_of_trucks, authority_status
        elif response.status_code == 403:
            p = multiprocessing.Process(target=play_sound_loop, args=("ah-shit-here-we-go-again.mp3",))
            p.start()
            input("press ENTER to stop playback")
            p.terminate()
            exit(0)
        else:
            print("Error: Failed to retrieve data")

        return None

    # below are the functions related to workbook/worksheet
    def __create_file_for_state_name__(self, state_name):
        self.__file_name__ = f"{state_name}.xlsx"
        if os.path.isfile(self.__file_name__):
            self.__workbook__ = openpyxl.load_workbook(filename=self.__file_name__)
            self.__lastSheet__ = self.__workbook__.worksheets[-1].title
            self.__worksheet__ = self.__workbook__.worksheets[-1]
        else:
            self.__workbook__ = openpyxl.Workbook()
            self.__workbook__.save(self.__file_name__)

    def __create_new_sheet__(self, sheet_name):
        if sheet_name not in self.__workbook__.sheetnames:
            self.__workbook__.create_sheet(sheet_name)
            self.__workbook__.active = self.__workbook__.sheetnames.index(sheet_name)
            self.__worksheet__ = self.__workbook__[sheet_name]
            self.__lastSheet__ = sheet_name
            self.__append_data_in_sheet__(data=self.__SHEET_HEADERS__)
            self.__workbook__.save(self.__file_name__)
        else:
            self.__worksheet__ = self.__workbook__[sheet_name]
            self.__lastSheet__ = sheet_name

    def __append_data_in_sheet__(self, data):
        if self.__workbook__ is None: 
            print("Error: workbook is None")
            exit(0)
        if self.__worksheet__ is None:
            print("Error: worksheet is None")
            return
        self.__worksheet__.append(data)
        self.__save_file__()

    def __save_file__(self):
        if self.__workbook__ is None:
            print("Error: workbook is None")
            return
        self.__workbook__.save(self.__file_name__)

    def run(self):
        state_name = input("Enter the state name: ").lower()
        # state_name="georgia"
        state_url = self.__check_if_company_exist__(state_name)

        if state_url is None:
            print(f"Company with name ${state_name} not found")
        else:
            self.__create_file_for_state_name__(state_name)
            list_of_all_cities_url = self.__get_all_cities__(state_url)

            to_skip = True if self.__lastSheet__ != None else False

            for list_of_cities_url in list_of_all_cities_url:
                for urls in list_of_cities_url:
                    sheet_name = urls.split('/')[-1]
                    if(self.__lastSheet__ == sheet_name):
                        to_skip = False
                    if to_skip:
                        print(f"Skipping {sheet_name}")
                        continue
                    self.__create_new_sheet__(sheet_name=sheet_name)
                    list_of_mc_numbers = self.__get_list_of_companiesVM__(urls)
                    print(f"{len(list_of_mc_numbers)} numbers of mc found in {urls}")

                    last_mc = self.__worksheet__[f"A{self.__worksheet__.max_row}"].value
                    print(f"Last MC number is {last_mc}")

                    mc_skip = True if last_mc != self.__SHEET_HEADERS__[0] else False

                    for mc_number in list_of_mc_numbers:
                        if mc_number == last_mc:
                            mc_skip = False
                        if mc_skip:
                            print(f"skip {mc_number}")
                            continue
                        # get safer data here
                        print(f"Starting request for {mc_number}.")
                        safer_data = self.__get_safer_data__(mcn=mc_number)
                        if safer_data:
                            (usdot, state_name, address, phone, email, number_of_trucks, authority_status) = safer_data
                            # append data in sheet
                            self.__append_data_in_sheet__(data=[mc_number, usdot, state_name, address, phone, email, number_of_trucks, authority_status])

                        print(f"Request for {mc_number} completed.")

                    print('**********************************************************')

# ##################################################################################################
#           NOTE:                                                                                  #
#           For this program to work flawlessly,                                                   #
#           Try not to create any sheet inside .xlsx manually                                      #
#           This behaviour can lead the program to skip every city until sheet name matches.       #
# ##################################################################################################
def main():
    runner = WebScrapper()
    runner.run()

if __name__ == "__main__":
    main()